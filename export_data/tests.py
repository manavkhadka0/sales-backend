from datetime import timedelta
from io import BytesIO

from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.test import APITestCase

from account.models import CustomUser, Factory, Franchise
from sales.models import Order


class ExportDataViewsTests(APITestCase):
    def setUp(self):
        # Create a salesperson
        self.sales_person = CustomUser.objects.create_user(
            username="salesperson",
            email="sales@example.com",
            password="password123",
            first_name="Sales",
            last_name="Person",
            role="SalesPerson",
        )

        now = timezone.now()
        # Older than 6 months (180 days)
        self.date_200_days_ago = now - timedelta(days=200)
        self.date_190_days_ago = now - timedelta(days=190)
        self.date_185_days_ago = now - timedelta(days=185)
        # Newer than 6 months
        self.date_5_days_ago = now - timedelta(days=5)

        # Create Factory
        self.factory = Factory.objects.create(name="Yachu Factory", short_form="yachu")

        # Create Franchise
        self.franchise1 = Franchise.objects.create(
            name="Franchise One", short_form="f1"
        )
        self.franchise2 = Franchise.objects.create(
            name="Franchise Two", short_form="f2"
        )

        # Create orders
        # 1. First unique old order (older than 6 months) -> John Doe
        self.order1 = Order.objects.create(
            full_name="John Doe",
            phone_number="1234567890",
            delivery_address="Address 1",
            payment_method="Cash on Delivery",
            sales_person=self.sales_person,
            created_at=self.date_200_days_ago,
            factory=self.factory,
            franchise=self.franchise1,
        )

        # 2. Duplicate old order (same name/phone, older than 6 months) -> John Doe
        self.order2 = Order.objects.create(
            full_name="John Doe",
            phone_number="1234567890",
            delivery_address="Address 2",
            payment_method="Cash on Delivery",
            sales_person=self.sales_person,
            created_at=self.date_190_days_ago,
            factory=self.factory,
            franchise=self.franchise1,
        )

        # 3. Second unique old order (older than 6 months) -> Jane Smith
        self.order3 = Order.objects.create(
            full_name="Jane Smith",
            phone_number="0987654321",
            delivery_address="Address 3",
            payment_method="Prepaid",
            sales_person=self.sales_person,
            created_at=self.date_190_days_ago,
            factory=self.factory,
            franchise=self.franchise2,
        )

        # 4. Duplicate old order (same name/phone, older than 6 months) -> Jane Smith
        self.order4 = Order.objects.create(
            full_name="Jane Smith",
            phone_number="0987654321",
            delivery_address="Address 4",
            payment_method="Prepaid",
            sales_person=self.sales_person,
            created_at=self.date_185_days_ago,
            factory=self.factory,
            franchise=self.franchise2,
        )

        # 5. New order (newer than 6 months) -> Alice Bob
        self.order5 = Order.objects.create(
            full_name="Alice Bob",
            phone_number="1112223333",
            delivery_address="Address 5",
            payment_method="Cash on Delivery",
            sales_person=self.sales_person,
            created_at=self.date_5_days_ago,
            factory=self.factory,
            franchise=self.franchise1,
        )

    def test_export_unique_old_orders(self):
        url = reverse("export-unique-old-orders")
        response = self.client.get(url)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        # Load Excel sheet and parse rows
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        # Check headers
        self.assertEqual(rows[0][3], "Customer Name")
        self.assertEqual(rows[0][4], "Contact Number")

        # We ordered by -id. In Django tests, IDs are autoincremented:
        # order1 (ID 1), order2 (ID 2), order3 (ID 3), order4 (ID 4)
        # Query iterator goes in descending order (-id):
        # order4 (ID 4) -> Jane Smith (0987654321) -> unique, added
        # order3 (ID 3) -> Jane Smith (0987654321) -> duplicate name/phone, skipped
        # order2 (ID 2) -> John Doe (1234567890) -> unique, added
        # order1 (ID 1) -> John Doe (1234567890) -> duplicate name/phone, skipped
        # So we expect 2 data rows: Jane Smith and John Doe
        self.assertEqual(len(rows), 3)  # Header + 2 data rows

        # Check customer names in the unique export
        names = {rows[1][3], rows[2][3]}
        self.assertEqual(names, {"Jane Smith", "John Doe"})

    def test_export_remaining_old_orders(self):
        url = reverse("export-remaining-old-orders")
        response = self.client.get(url)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        # Load Excel sheet and parse rows
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        # We expect remaining orders to be:
        # Those older than 6 months EXCEPT the unique ones.
        # Unique ones selected by export-unique-old-orders are order4 and order2.
        # Remaining ones should be order3 and order1.
        # Note that order5 is newer than 6 months, so it shouldn't be included anywhere.
        self.assertEqual(len(rows), 3)  # Header + 2 data rows

        # Order of export: sorted by -id:
        # order3 (ID 3) -> Jane Smith (Address 3)
        # order1 (ID 1) -> John Doe (Address 1)
        self.assertEqual(rows[1][3], "Jane Smith")
        self.assertEqual(rows[1][6], "Address 3")
        self.assertEqual(rows[2][3], "John Doe")
        self.assertEqual(rows[2][6], "Address 1")

    def test_yachu_full_export_filtering(self):
        url = reverse("yachu-full-export")

        # 1. No filters: returns all 5 orders
        response = self.client.get(url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn("text/csv", response["Content-Type"])

        content = b"".join(response.streaming_content).decode("utf-8-sig")
        lines = [line for line in content.strip().split("\r\n") if line]
        # Header + 5 orders
        self.assertEqual(len(lines), 6)

        # 2. Filter by start_date (e.g. date_190_days_ago)
        # Should include orders created at: 190 days ago, 185 days ago, and 5 days ago (orders 2, 3, 4, 5)
        start_date_str = self.date_190_days_ago.strftime("%Y-%m-%d")
        response = self.client.get(url, {"start_date": start_date_str})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        content = b"".join(response.streaming_content).decode("utf-8-sig")
        lines = [line for line in content.strip().split("\r\n") if line]
        self.assertEqual(len(lines), 5)  # Header + 4 orders

        # 3. Filter by end_date (e.g. date_185_days_ago)
        # Should include orders created at: 200 days ago, 190 days ago, 185 days ago (orders 1, 2, 3, 4)
        end_date_str = self.date_185_days_ago.strftime("%Y-%m-%d")
        response = self.client.get(url, {"end_date": end_date_str})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        content = b"".join(response.streaming_content).decode("utf-8-sig")
        lines = [line for line in content.strip().split("\r\n") if line]
        self.assertEqual(len(lines), 5)  # Header + 4 orders

        # 4. Filter by both start_date and end_date
        # Should include orders created between 190 days ago and 185 days ago (orders 2, 3, 4)
        response = self.client.get(
            url, {"start_date": start_date_str, "end_date": end_date_str}
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        content = b"".join(response.streaming_content).decode("utf-8-sig")
        lines = [line for line in content.strip().split("\r\n") if line]
        self.assertEqual(len(lines), 4)  # Header + 3 orders

        # 5. Filter by franchise (ID)
        # Franchise One should have orders 1, 2, 5
        response = self.client.get(url, {"franchise": self.franchise1.id})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        content = b"".join(response.streaming_content).decode("utf-8-sig")
        lines = [line for line in content.strip().split("\r\n") if line]
        self.assertEqual(len(lines), 4)  # Header + 3 orders

        # 6. Filter by franchise (Name)
        # Franchise Two should have orders 3, 4
        response = self.client.get(url, {"franchise": "Franchise Two"})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        content = b"".join(response.streaming_content).decode("utf-8-sig")
        lines = [line for line in content.strip().split("\r\n") if line]
        self.assertEqual(len(lines), 3)  # Header + 2 orders

        # 7. Invalid date format
        response = self.client.get(url, {"start_date": "invalid-date"})
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
